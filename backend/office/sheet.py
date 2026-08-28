"""Hesap tablosu — gerçek .xlsx, gerekçeli değişikliklerle.

Kendi biçimimizi uydurmuyoruz. Dosya gerçek bir `.xlsx`; Berkay onu birine
yollayabilir, karşı taraf Excel'de açar. Fark dosya biçiminde değil, ajanın
ona nasıl eriştiğinde: bir arayüzü sürmek yerine hücre modeline doğrudan
yazıyor ve her yazma neden yapıldığını taşıyor.

Formül hesaplama henüz yok — formüller dosyaya doğru yazılıyor ve Excel'de
çalışıyor ama biz değerlerini göremiyoruz. `formulas` motoru bir sonraki
adım; `read` çağrısı bu boşluğu sessizce geçmiyor, açıkça söylüyor.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

from .calc import CalcError, evaluate
from .model import OfficeDocument

#: Tek okumada modele gidecek en fazla hücre. Bir tabloyu bağlama komple
#: dökmek ajanın geri kalan işi için yer bırakmıyor.
MAX_CELLS = 4000


class SheetError(RuntimeError):
    """Tablo işlemi yapılamadı."""


@dataclass
class Workbook(OfficeDocument):
    book: openpyxl.Workbook = field(default_factory=openpyxl.Workbook)

    #: (defter uzunluğu, sonuçlar) — tablo değişmediyse yeniden hesaplanmaz.
    _calc_cache: tuple[int, dict[str, str]] | None = field(
        default=None, repr=False, compare=False
    )

    @property
    def kind(self) -> str:
        return "tablo"

    def values(self) -> dict[str, str]:
        """Formüllerin hesaplanmış sonuçları. Hesaplanamazsa boş sözlük.

        Ajan formül sonucunu kendi aritmetiğiyle tahmin etmemeli; bu
        sözlük onun yerine geçiyor.
        """
        stamp = len(self.ledger)
        if self._calc_cache is not None and self._calc_cache[0] == stamp:
            return self._calc_cache[1]
        has_formula = any(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for sheet in self.book.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        if not has_formula:
            result: dict[str, str] = {}
        else:
            try:
                result = evaluate(self.book, list(self.book.sheetnames))
            except CalcError:
                result = {}
        self._calc_cache = (stamp, result)
        return result

    @classmethod
    def create(cls, path: str, sheet_name: str = "Sheet1") -> Workbook:
        book = openpyxl.Workbook()
        book.active.title = sheet_name
        return cls(path=path, book=book)

    @classmethod
    def open(cls, path: str) -> Workbook:
        target = Path(path)
        if not target.exists():
            raise SheetError(f"{target} yok")
        try:
            book = openpyxl.load_workbook(target)
        except Exception as exc:
            raise SheetError(f"{target} açılamadı: {exc}") from None
        return cls(path=str(target), book=book)

    # --- gezinme ----------------------------------------------------------

    def _sheet(self, name: str | None):
        if name is None:
            return self.book.active
        if name not in self.book.sheetnames:
            raise SheetError(
                f"{name!r} sayfası yok. Sayfalar: {', '.join(self.book.sheetnames)}"
            )
        return self.book[name]

    def summary(self) -> str:
        lines = [f"{self.path} ({self.kind})"]
        for ws in self.book.worksheets:
            used = f"{get_column_letter(ws.max_column)}{ws.max_row}"
            lines.append(f"  {ws.title}: A1:{used} ({ws.max_row} satır, {ws.max_column} sütun)")
        if self.ledger.dirty:
            lines.append(f"  {self.ledger.unsaved_count} kaydedilmemiş değişiklik")
        return "\n".join(lines)

    def add_sheet(self, name: str, why: str) -> str:
        if name in self.book.sheetnames:
            raise SheetError(f"{name!r} sayfası zaten var")
        self.book.create_sheet(name)
        self.ledger.record(f"sayfa {name}", None, "oluşturuldu", why)
        return f"{name!r} sayfası eklendi."

    # --- okuma ------------------------------------------------------------

    def read(self, ref: str | None = None, sheet: str | None = None) -> str:
        ws = self._sheet(sheet)
        if ref:
            bounds = _parse_range(ref)
        else:
            bounds = (1, 1, ws.max_column or 1, ws.max_row or 1)

        min_col, min_row, max_col, max_row = bounds
        cells = (max_col - min_col + 1) * (max_row - min_row + 1)
        truncated = False
        if cells > MAX_CELLS:
            max_row = min_row + max(1, MAX_CELLS // max(1, max_col - min_col + 1)) - 1
            truncated = True

        computed = self.values()
        lines = []
        formulas_seen = 0
        uncomputed = 0
        for row in range(min_row, max_row + 1):
            values = []
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas_seen += 1
                    result = computed.get(f"{ws.title}!{cell.coordinate}")
                    if result is None:
                        uncomputed += 1
                        values.append(value)
                    else:
                        # Hem formül hem sonuç: model neyin yazdığını da
                        # neyin çıktığını da görmeli.
                        values.append(f"{value} → {result}")
                    continue
                values.append("" if value is None else str(value))
            lines.append(f"{row}\t" + "\t".join(values))

        header = "\t" + "\t".join(
            get_column_letter(c) for c in range(min_col, max_col + 1)
        )
        out = [f"{ws.title}!{_range_label(min_col, min_row, max_col, max_row)}", header, *lines]
        if truncated:
            out.append(f"[{MAX_CELLS} hücrede kesildi — daha dar bir aralık iste]")
        if uncomputed:
            # Bu boşluğu gizlemek, modelin formül sonucunu bildiğini sanmasına
            # yol açar ve o varsayımla yanlış rapor yazar.
            out.append(
                f"[{uncomputed} formül hesaplanamadı; sonucunu tahmin etme, "
                f"kullanıcıya hesaplanamadığını söyle]"
            )
        elif formulas_seen:
            out.append(
                f"[{formulas_seen} formülün sonucu ok işaretinden sonra "
                f"yazılı; sonucu kendin hesaplama, oradakini kullan]"
            )
        return "\n".join(out)

    # --- yazma ------------------------------------------------------------

    def write(self, ref: str, values: list[list[Any]], why: str,
              sheet: str | None = None) -> str:
        ws = self._sheet(sheet)
        if not values or not isinstance(values, list):
            raise SheetError("values boş olamaz; satır listesi bekleniyor")

        min_col, min_row, _max_col, _max_row = _parse_range(ref)
        written = 0
        for row_offset, row_values in enumerate(values):
            if not isinstance(row_values, list):
                row_values = [row_values]
            for col_offset, value in enumerate(row_values):
                cell = ws.cell(row=min_row + row_offset, column=min_col + col_offset)
                before = cell.value
                if before == value:
                    continue
                cell.value = value
                self.ledger.record(
                    f"{ws.title}!{cell.coordinate}", before, value, why
                )
                written += 1

        if not written:
            return "Değer zaten aynıydı, hiçbir hücre değişmedi."
        return f"{written} hücre yazıldı ({ws.title}!{ref}) — {why}"

    # --- kaydetme ve geri alma --------------------------------------------

    def save(self, path: str | None = None) -> str:
        target = Path(path or self.path)
        target.parent.mkdir(parents=True, exist_ok=True)
        try:
            self.book.save(target)
        except Exception as exc:
            raise SheetError(f"{target} kaydedilemedi: {exc}") from None
        self.path = str(target)
        self.ledger.mark_saved()
        return f"{target} kaydedildi ({len(self.ledger)} değişiklik)."

    def undo(self, count: int = 1) -> str:
        changes = self.ledger.last(count)
        if not changes:
            return "Geri alınacak değişiklik yok."

        undone = 0
        for change in changes:
            match = re.fullmatch(r"(.+)!([A-Z]+\d+)", change.target)
            if not match:
                break  # hücre olmayan değişiklikler (sayfa ekleme) geri alınmıyor
            sheet_name, coordinate = match.groups()
            if sheet_name not in self.book.sheetnames:
                break
            self.book[sheet_name][coordinate].value = change.before
            undone += 1

        self.ledger.drop_last(undone)
        if undone < len(changes):
            return (
                f"{undone} değişiklik geri alındı. Kalanlar hücre değişikliği "
                f"olmadığı için geri alınamadı."
            )
        return f"{undone} değişiklik geri alındı."


def _parse_range(ref: str) -> tuple[int, int, int, int]:
    """`"B2"` ya da `"A1:C10"` -> (min_col, min_row, max_col, max_row)."""
    cleaned = ref.strip().upper().replace("$", "")
    if "!" in cleaned:
        cleaned = cleaned.split("!", 1)[1]
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cleaned)
    except Exception:
        raise SheetError(
            f"{ref!r} geçerli bir hücre aralığı değil. Örnek: A1, B2:D10"
        ) from None
    if None in (min_col, min_row, max_col, max_row):
        raise SheetError(f"{ref!r} sınırsız aralık; satır numarası ver (A1:A100 gibi)")
    return min_col, min_row, max_col, max_row


def _range_label(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"
